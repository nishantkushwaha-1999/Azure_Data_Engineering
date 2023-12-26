# Databricks notebook source
!pip install openpyxl

# COMMAND ----------

import pandas as pd
import os
import shutil
import warnings
import openpyxl as opxl
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill, Color

# COMMAND ----------

Flights_pre_prod = pd.read_csv('dbfs:/mnt/xxairlinesprod/flights"')
Flights_test = pd.read_csv('/dbfs/FileStore/shared_uploads/nskp.sec19@gmail.com/Test_Data/flights.csv')   

# COMMAND ----------

records_pre_prod = len(Flights_pre_prod)
records_test = len(Flights_test)
print(records_pre_prod)
print(records_test)

# COMMAND ----------

eng_fields = set(eng.columns)
cdp_fields = set(cdp.columns)
result = list(eng_fields.intersection(cdp_fields))
result.sort()
print(eng_fields-cdp_fields)

# COMMAND ----------

comparison = []
Flights_pre_prod_filtered = Flights_pre_prod[result]
Flights_test_filtered = Flights_test[result]
for item in result:
    i = []
    a = str(item)
    i.append(a)
    Flights_pre_prod_count_null = Flights_pre_prod[a].isnull().sum()
    Flights_pre_prod_fill_rate = round(((records_Flights_pre_prod-Flights_pre_prod_count_null)/records_Flights_pre_prod)*100,2)
    Flights_test_count_null = Flights_test[a].isnull().sum()
    Flights_test_fill_rate = round(((records_Flights_test-Flights_test_count_null)/records_Flights_test)*100,2)
    i.append(Flights_pre_prod_count_null)
    i.append(Flights_test_count_null)
    i.append(records_Flights_pre_prod-Flights_pre_prod_count_null)
    i.append(records_Flights_test-Flights_test_count_null)
    i.append(Flights_pre_prod_fill_rate)
    i.append(Flights_test_fill_rate)
    uniques_Flights_pre_prod = list(Flights_pre_prod[a].unique())
    uniques_Flights_test = list(Flights_test[a].unique())
    
    if (len(uniques_Flights_pre_prod) <= 10) and (len(uniques_Flights_test) <= 10) and ('Date' not in a) and ('Day' not in a):
        unique_count_Flights_pre_prod = []
        unique_count_Flights_test = []
        for unique_Flights_pre_prod in uniques_Flights_pre_prod:
            if str(unique_Flights_pre_prod) == 'nan':
                count_unique = Flights_pre_prod[a].isnull().sum()
            else:
                count_unique = len(Flights_pre_prod[a].loc[Flights_pre_prod[a]==unique_Flights_pre_prod])
            unique_count_Flights_pre_prod.append(str(unique_Flights_pre_prod)+' '+str(count_unique))
        for unique_Flights_test in uniques_Flights_test:
            if str(unique_Flights_test) == 'nan':
                count_unique = Flights_test[a].isnull().sum()
            else:
                count_unique = len(Flights_test[a].loc[Flights_test[a]==unique_Flights_test])
            unique_count_Flights_test.append(str(unique_Flights_test)+' '+str(count_unique))
        unique_count_Flights_pre_prod.sort()
        unique_count_Flights_test.sort()
        unique_count_Flights_pre_prod = '|'.join(unique_count_Flights_pre_prod)
        unique_count_Flights_test = '|'.join(unique_count_Flights_test)
        unique_count_Flights_pre_prod = unique_count_Flights_pre_prod.replace('nan', 'null')
        unique_count_Flights_test = unique_count_Flights_test.replace('nan', 'null')
        i.append(unique_count_Flights_pre_prod)
        i.append(unique_count_Flights_test)
    else:
        i.append('')
        i.append('')
    
    # Fill in the Unique Id
    base = ''
    if a != base and ('Date' not in a) and ('Day' not in a):
        Flights_pre_prod_item = Flights_pre_prod[[base, a]]
        Flights_test_item = Flights_test[[base, a]]
        if len(Flights_pre_prod_item) >=  len(Flights_test_item):
            temp = Flights_pre_prod_item.merge(Flights_test_item, on=base, how='left')
        else:
            temp = Flights_test_item.merge(Flights_pre_prod_item, on=base, how='left')
        
        try:
            temp[a+str('_x')] = temp[a+str('_x')].apply(str.upper)
            temp[a+str('_y')] = temp[a+str('_y')].apply(str.upper)
        except TypeError:
            pass
        temp['sim'] = temp[[a+str('_x'), a+str('_y')]].apply(tuple, axis=1)
        # print(temp)
        
        count = 0
        for row in temp['sim']:
            Flights_pre_prod_row, Flights_test_row = row
            if str(Flights_pre_prod_row) == str(Flights_test_row):
                count += 1
        total_count = max(records_Flights_pre_prod, records_Flights_test)
        similarity = round((count/total_count)*100, 2)
        # print(a, similarity)
        i.append(similarity)
    elif a == base:
        Flights_pre_prod_item = list(Flights_pre_prod[a])
        Flights_test_item = list(Flights_test[a])
        if len(Flights_pre_prod_item) >= len(Flights_test_item):
            common_item = [i for i in Flights_pre_prod_item if i in Flights_test_item]
        else:
            common_item = [i for i in Flights_test_item if i in Flights_pre_prod_item]
        total_count = max(records_Flights_pre_prod, records_Flights_test)
        similarity = round((len(common_item)/total_count)*100, 2)
        i.append(similarity)
    else:
        i.append('')
    
    comparison.append(i)
# print(comparison)

comparison = pd.DataFrame(comparison, columns =['Flights_test Field', 'Flights_pre_prod Null Count','Analytics Null Count','Flights_pre_prod Non Null Count','Analytics Non Null Count','Flights_pre_prod Fill Rate',
                                               'Analytics Fill Rate','Flights_pre_prod Unique Count','Analytics Unique Count', 'Similarity']) 

comparison['Notes'] = ''
# comparison

# COMMAND ----------

description = [['','',''],
                ['', f'*The file was shared by the team on 02/01/2023 and the file was validated on 02/01/2023',''],	
                ['',f'Record counts in the engineering shared file was {records_pre_prod}',''],
                ['',f'Record counts in the analytics created file was {records_test}',''], 	
                ['', '',''],
                ['', 'Notes', 'Context behind a lower similarity score or different fill rate']]

description = pd.DataFrame(description, columns=['','',''])

# COMMAND ----------

comparison['Similarity'] = comparison['Similarity'].replace('', -1)
comparison = comparison.astype({'Similarity': float})
re_eval = comparison.loc[(comparison['Eng Fill Rate'] != comparison['Analytics Fill Rate']) | ((comparison['Similarity'] < 100.0) & (comparison['Similarity'] > 0.0))]
comparison['Similarity'] = comparison['Similarity'].replace(-1, '')

# COMMAND ----------

index_for_fill = re_eval.index.to_list()

name = 'validation_summary_ISV_CDP.xlsx'
files_sys = '/dbfs/FileStore'
adb_instance = 'https://adb-7585872976963873.13.azuredatabricks.net/files/'
path = '/shared_uploads/v-kushwahan@microsoft.com/Validation_Reports/'

with pd.ExcelWriter(name, engine='openpyxl') as writer:
    description.to_excel(writer, sheet_name="Summary Glossary", index=False, header=False)
    ws = writer.sheets['Summary Glossary']
    ws['B2'].font = Font(bold=True)
    ws['B3'].font = Font(bold=True)
    ws['B4'].font = Font(bold=True)
    ws['B5'].font = Font(bold=True)
    ws['B7'].font = Font(bold=True)
    ws['C7'].font = Font(bold=True)
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 24.50
    ws.column_dimensions['C'].width = 119.30

    main_border = Side(border_style='thin',color="000000")
    inner_border = Side(border_style='thin',color="D3D3D3")
    ws['B7'].border = Border(top=main_border,left=main_border, right=inner_border, bottom=inner_border)
    ws['B18'].border = Border(top=inner_border,left=main_border, right=inner_border, bottom=main_border)
    ws['C7'].border = Border(top=main_border,left=inner_border, right=main_border, bottom=inner_border)
    ws['C18'].border = Border(top=inner_border,left=inner_border, right=main_border, bottom=main_border)

    range = ws['B8':'B17']
    for cell in range:
        for x in cell:
            x.border = Border(top=inner_border,left=main_border, right=inner_border, bottom=inner_border)

    range = ws['C8':'C17']
    for cell in range:
        for x in cell:
            x.border = Border(top=inner_border, left=inner_border, right=main_border, bottom=inner_border)

    range = ws['B7':'C18']
    for cell in range:
        for x in cell:
            x.alignment = Alignment(horizontal='center', vertical='center')

    ws.sheet_view.showGridLines = False
    ws.sheet_view.showRowColHeaders = False
    
    comparison.to_excel(writer, sheet_name="Validation Summary", index=False)
    ws = writer.sheets['Validation Summary']
    ws.column_dimensions['A'].width = 45.14
    ws.column_dimensions['B'].width = 13.43
    ws.column_dimensions['C'].width = 18.43
    ws.column_dimensions['D'].width = 17.71
    ws.column_dimensions['E'].width = 22.71
    ws.column_dimensions['F'].width = 11
    ws.column_dimensions['G'].width = 16
    ws.column_dimensions['H'].width = 197.29
    ws.column_dimensions['I'].width = 197.29
    ws.column_dimensions['J'].width = 8.71
    ws.column_dimensions['K'].width = 100

    for row in index_for_fill:
        count = 1
        while count != 12:
            ws.cell(row=row+2, column=count).fill = PatternFill(fgColor=Color('FFFF00'), patternType='solid')
            count += 1

shutil.copy2(name, files_sys + path + name)
os.remove(name)
print("The report is accessible at: ", adb_instance+path+name)
